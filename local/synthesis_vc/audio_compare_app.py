from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, unquote, urlparse
import argparse
import csv
import json
import mimetypes
from openpyxl import load_workbook


ROOT = Path.cwd().resolve()
OUT = ROOT / "outputs"
BASE_VERSION = OUT / "openvoice-v2-20voices"
TRANSLATION_XLSX = OUT / "mms-tts-zyb" / "壮语文本校对与译意.xlsx"


def safe_path(value):
    path = (ROOT / unquote(value)).resolve()
    if OUT not in path.parents and path != OUT:
        raise ValueError("path outside outputs")
    return path


def version_label(path):
    labels = {
        "openvoice-v2-20voices": "旧版转换",
        "openvoice-v2-20voices-improved": "缓解失真版",
        "openvoice-v2-20voices-improved-norm": "参考归一化版",
    }
    return labels.get(path.name, path.name.replace("openvoice-v2-", ""))


def versions():
    found = []
    for path in sorted(OUT.glob("openvoice-v2-*")):
        if path.name.endswith("-test"):
            continue
        if (path / "manifest.csv").exists() and (path / "selected_speakers.csv").exists():
            found.append(path)
    return found


def selected_refs(version):
    refs = {}
    with (version / "selected_speakers.csv").open(encoding="utf-8") as f:
        for row in csv.DictReader(f):
            refs[row["speaker_id"]] = row.get("demo_reference_wav") or row.get("reference_wav")
    return refs


def translations():
    if not TRANSLATION_XLSX.exists():
        return {}
    wb = load_workbook(TRANSLATION_XLSX, read_only=True, data_only=True)
    ws = wb.active
    header = [c for c in next(ws.iter_rows(values_only=True))]
    text_i = header.index("原文")
    translation_i = header.index("中文译意/摘要")
    return {
        row[text_i]: row[translation_i]
        for row in ws.iter_rows(values_only=True)
        if row[text_i] and row[translation_i]
    }


def load_items():
    all_versions = versions()
    refs = selected_refs(all_versions[-1])
    zh_by_text = translations()
    by_version = {}
    text_by_key = {}
    for version in all_versions:
        rows = {}
        with (version / "manifest.csv").open(encoding="utf-8") as f:
            for row in csv.DictReader(f):
                src = Path(row["source_wav"]).name
                key = (row["speaker_id"], src)
                rows[key] = str((ROOT / row["converted_wav"]).relative_to(ROOT))
                if row.get("original_text"):
                    text_by_key[key] = row["original_text"]
        by_version[version.name] = rows

    items = []
    seen = set()
    with (BASE_VERSION / "manifest.csv").open(encoding="utf-8") as f:
        for row in csv.DictReader(f):
            speaker = row["speaker_id"]
            source = row["source_wav"]
            source_name = Path(source).name
            key = (speaker, source_name)
            if key in seen:
                continue
            seen.add(key)
            src = ROOT / source
            ref = ROOT / refs[speaker]
            conversions = []
            for version in all_versions:
                converted = by_version.get(version.name, {}).get(key)
                if converted and (ROOT / converted).exists():
                    conversions.append({"label": version_label(version), "path": converted})
            if not (src.exists() and ref.exists() and conversions):
                continue
            text = row.get("original_text") or text_by_key.get(key, "")
            items.append({
                "id": len(items),
                "label": f"{Path(source).stem} / {speaker}",
                "speaker": speaker,
                "text": text,
                "translation": zh_by_text.get(text, ""),
                "source": source,
                "reference": str(ref.relative_to(ROOT)),
                "conversions": conversions,
            })
    return items


def audio_url(path):
    return "/audio?path=" + quote(path)


class Handler(BaseHTTPRequestHandler):
    def send(self, code, body, content_type):
        raw = body.encode("utf-8") if isinstance(body, str) else body
        self.send_response(code)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(raw)))
        self.end_headers()
        self.wfile.write(raw)

    def head(self, code, size, content_type):
        self.send_response(code)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(size))
        self.end_headers()

    def do_HEAD(self):
        parsed = urlparse(self.path)
        if parsed.path in ("/", "/index.html"):
            self.head(200, len(HTML.encode("utf-8")), "text/html; charset=utf-8")
            return
        if parsed.path == "/audio":
            try:
                path = safe_path(parse_qs(parsed.query)["path"][0])
                size = path.stat().st_size
            except Exception:
                self.head(404, 0, "text/plain; charset=utf-8")
                return
            self.head(200, size, mimetypes.guess_type(path.name)[0] or "audio/wav")
            return
        self.head(404, 0, "text/plain; charset=utf-8")

    def do_GET(self):
        parsed = urlparse(self.path)
        if parsed.path in ("/", "/index.html"):
            self.send(200, HTML, "text/html; charset=utf-8")
            return
        if parsed.path == "/api/items":
            self.send(200, json.dumps(load_items(), ensure_ascii=False), "application/json; charset=utf-8")
            return
        if parsed.path == "/audio":
            try:
                path = safe_path(parse_qs(parsed.query)["path"][0])
                data = path.read_bytes()
            except Exception:
                self.send(404, "not found", "text/plain; charset=utf-8")
                return
            self.send(200, data, mimetypes.guess_type(path.name)[0] or "audio/wav")
            return
        self.send(404, "not found", "text/plain; charset=utf-8")


HTML = r"""<!doctype html>
<html lang="zh-CN">
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>壮语 TTS 音色转换对比</title>
<style>
body{margin:0;font:14px/1.5 system-ui,-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;background:#f6f7f9;color:#14171f}
main{max-width:1180px;margin:auto;padding:18px}
header{display:flex;gap:12px;align-items:center;justify-content:space-between;margin-bottom:10px}
h1{font-size:22px;margin:0}
select{width:min(780px,100%);padding:8px;border:1px solid #cfd5df;border-radius:6px;background:#fff}
.text{white-space:pre-wrap;background:#fff;border:1px solid #e1e5eb;border-radius:6px;padding:10px;max-height:110px;overflow:auto}
.grid{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-top:10px}
.full{grid-column:1/-1}.panel{background:#fff;border:1px solid #e1e5eb;border-radius:6px;padding:10px}
.panel h2{font-size:14px;margin:0 0 6px} audio{width:100%;height:34px}
canvas{width:100%;height:68px;background:#fbfcfe;border:1px solid #edf0f5;border-radius:4px;margin-top:6px}
@media(max-width:760px){.grid{grid-template-columns:1fr}header{display:block}select{margin-top:10px}}
</style>
<main>
  <header><h1>壮语 TTS 音色转换对比</h1><select id="pick"></select></header>
  <section class="text"><b>原文</b><div id="text"></div><br><b>中文译意/摘要</b><div id="translation"></div></section>
  <section class="grid" id="grid">
    <div class="panel full"><h2>mms-tts-zyb 合成语音</h2><audio controls id="source"></audio><canvas id="sourceWave"></canvas></div>
    <div class="panel"><h2>参考语音</h2><audio controls id="reference"></audio><canvas id="referenceWave"></canvas></div>
  </section>
</main>
<script>
let data=[];
fetch("/api/items").then(r=>r.json()).then(items=>{data=items;init();});
function audioUrl(path){return "/audio?path="+encodeURIComponent(path);}
function init(){
  data.forEach((x,i)=>{const o=document.createElement("option");o.value=i;o.textContent=x.label;pick.appendChild(o);});
  pick.onchange=()=>show(data[pick.value]);
  show(data[0]);
}
function show(x){
  text.textContent=x.text||"无原文";
  translation.textContent=x.translation||"无译意";
  setAudio("source", x.source);
  setAudio("reference", x.reference);
  document.querySelectorAll(".conversion").forEach(el=>el.remove());
  x.conversions.forEach((c,i)=>{
    const id="conversion"+i;
    const panel=document.createElement("div");
    panel.className="panel conversion";
    panel.innerHTML=`<h2>${c.label}</h2><audio controls id="${id}"></audio><canvas id="${id}Wave"></canvas>`;
    grid.appendChild(panel);
    setAudio(id, c.path);
  });
}
function setAudio(id,path){
  const url=audioUrl(path);
  document.getElementById(id).src=url;
  drawWave(url, document.getElementById(id+"Wave"));
}
async function drawWave(url, canvas){
  const ctx=canvas.getContext("2d"), dpr=devicePixelRatio||1;
  canvas.width=canvas.clientWidth*dpr; canvas.height=canvas.clientHeight*dpr; ctx.clearRect(0,0,canvas.width,canvas.height);
  try{
    const buf=await fetch(url).then(r=>r.arrayBuffer());
    const ac=new AudioContext(); const audio=await ac.decodeAudioData(buf); await ac.close();
    const ch=audio.getChannelData(0), w=canvas.width, h=canvas.height, mid=h/2, step=Math.ceil(ch.length/w);
    ctx.strokeStyle="#2b6cb0"; ctx.beginPath();
    for(let x=0;x<w;x++){let min=1,max=-1;for(let i=x*step;i<Math.min((x+1)*step,ch.length);i++){const v=ch[i]; if(v<min)min=v; if(v>max)max=v;}ctx.moveTo(x,mid+min*mid*.9);ctx.lineTo(x,mid+max*mid*.9);}
    ctx.stroke();
  }catch(e){ctx.fillText("波形加载失败", 10, 24);}
}
</script>
</html>
"""


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--host", default="0.0.0.0")
    parser.add_argument("--port", type=int, default=8787)
    args = parser.parse_args()
    ThreadingHTTPServer((args.host, args.port), Handler).serve_forever()


if __name__ == "__main__":
    main()
