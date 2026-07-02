from pathlib import Path
import argparse
import csv
import shutil
import tempfile

import numpy as np
import soundfile as sf
import torch
from openpyxl import load_workbook
from openvoice.api import ToneColorConverter


AISHELL = Path("/data/duhu/dbase/aishell/data_aishell/wav/train")
SRC_MANIFEST = Path("outputs/mms-tts-zyb/manifest.xlsx")
MODEL_DIR = Path.home() / ".cache/huggingface/hub/models--myshell-ai--OpenVoiceV2/snapshots/f36e7edfe1684461a8343844af60babc2efbb727"
OUT = Path("outputs/openvoice-v2-20voices")


class _NoWatermark:
    def to(self, device):
        return self


def wav_seconds(path: Path) -> float:
    info = sf.info(path)
    return info.frames / info.samplerate


def normalize_wav(src: Path, dst: Path, peak=0.8):
    audio, sr = sf.read(src, dtype="float32")
    current = float(np.max(np.abs(audio))) if audio.size else 0.0
    if current > 0:
        audio = audio * (peak / current)
    sf.write(dst, audio, sr)


def demo():
    with tempfile.TemporaryDirectory() as d:
        src = Path(d) / "in.wav"
        dst = Path(d) / "out.wav"
        sf.write(src, np.array([0.0, 0.1, -0.2], dtype=np.float32), 16000)
        normalize_wav(src, dst, 0.8)
        y, _ = sf.read(dst, dtype="float32")
        assert abs(float(np.max(np.abs(y))) - 0.8) < 1e-4


def pick_refs(n=20, refs_per_speaker=1):
    refs = []
    for spk in sorted(p for p in AISHELL.iterdir() if p.is_dir()):
        wavs = sorted(spk.glob("*.wav"))
        good = [w for w in wavs if 3 <= wav_seconds(w) <= 12]
        if good:
            mid = len(good) // 2
            refs.append((spk.name, good[mid:mid + refs_per_speaker] or good[-refs_per_speaker:]))
        if len(refs) == n:
            return refs
    raise RuntimeError(f"only found {len(refs)} usable reference speakers")


def load_sources():
    wb = load_workbook(SRC_MANIFEST, read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))[1:]
    return [(f"{r[0]}_r{r[1]}_c{r[2]}_{r[3]}", Path(r[4]), r[5], r[6]) for r in rows if r and r[4]]


def split_wav(path, seconds, tmpdir):
    audio, sr = sf.read(path, dtype="float32")
    if audio.ndim > 1:
        audio = audio.mean(axis=1)
    chunk = int(sr * seconds)
    if len(audio) <= chunk:
        return [path]
    paths = []
    for i, start in enumerate(range(0, len(audio), chunk)):
        part = audio[start:start + chunk]
        if len(part) < sr * 0.2:
            continue
        out = Path(tmpdir) / f"{path.stem}_{i:03d}.wav"
        sf.write(out, part, sr)
        paths.append(out)
    return paths


def convert_segmented(vc, src, src_se, tgt_se, dst, tau, segment_seconds):
    if not segment_seconds or wav_seconds(src) <= segment_seconds:
        vc.convert(str(src), src_se=src_se, tgt_se=tgt_se, output_path=str(dst), tau=tau)
        return
    with tempfile.TemporaryDirectory() as d:
        audios = []
        sr = None
        for part in split_wav(src, segment_seconds, d):
            part_se = vc.extract_se(str(part)) if part != src else src_se
            audio = vc.convert(str(part), src_se=part_se, tgt_se=tgt_se, tau=tau)
            audios.append(audio)
            sr = vc.hps.data.sampling_rate
        sf.write(dst, np.concatenate(audios), sr)


def main(limit_sources=None, limit_speakers=None, out=OUT, refs_per_speaker=1, tau=0.3, segment_seconds=None, normalize_refs_peak=None):
    device = "cuda:7" if torch.cuda.is_available() else "cpu"
    import wavmark
    wavmark.load_model = lambda: _NoWatermark()
    vc = ToneColorConverter(str(MODEL_DIR / "converter/config.json"), device=device)
    vc.watermark_model = None
    vc.load_ckpt(str(MODEL_DIR / "converter/checkpoint.pth"))

    refs = pick_refs(limit_speakers or 20, refs_per_speaker)
    sources = load_sources()[:limit_sources]
    out = Path(out)
    (out / "speaker_embeddings").mkdir(parents=True, exist_ok=True)
    (out / "source_embeddings").mkdir(parents=True, exist_ok=True)
    (out / "refs").mkdir(parents=True, exist_ok=True)

    target_ses = {}
    with (out / "selected_speakers.csv").open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["speaker_id", "reference_wavs", "demo_reference_wav", "duration_sec"])
        for spk, ref_list in refs:
            se_path = out / "speaker_embeddings" / f"{spk}.pth"
            demo_ref = out / "refs" / f"{spk}.wav"
            refs_for_se = ref_list
            if normalize_refs_peak:
                norm_dir = out / "normalized_refs" / spk
                norm_dir.mkdir(parents=True, exist_ok=True)
                refs_for_se = []
                for ref in ref_list:
                    norm_ref = norm_dir / ref.name
                    if not norm_ref.exists():
                        normalize_wav(ref, norm_ref, normalize_refs_peak)
                    refs_for_se.append(norm_ref)
            if not demo_ref.exists():
                shutil.copyfile(refs_for_se[0], demo_ref)
            target_ses[spk] = torch.load(se_path, map_location=device).to(device) if se_path.exists() else vc.extract_se([str(p) for p in refs_for_se], str(se_path))
            writer.writerow([spk, ";".join(map(str, ref_list)), demo_ref, round(sum(wav_seconds(p) for p in ref_list), 3)])

    with (out / "manifest.csv").open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["speaker_id", "source_id", "source_wav", "converted_wav", "original_text", "tts_text"])
        for source_id, src, original_text, tts_text in sources:
            src_se_path = out / "source_embeddings" / f"{source_id}.pth"
            src_se = torch.load(src_se_path, map_location=device).to(device) if src_se_path.exists() else vc.extract_se(str(src), str(src_se_path))
            for spk, _ in refs:
                dst = out / spk / src.name
                dst.parent.mkdir(parents=True, exist_ok=True)
                if not dst.exists():
                    convert_segmented(vc, src, src_se, target_ses[spk], dst, tau, segment_seconds)
                writer.writerow([spk, source_id, src, dst, original_text, tts_text])
                print(dst)


if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--out", default=str(OUT))
    p.add_argument("--limit-sources", type=int)
    p.add_argument("--limit-speakers", type=int)
    p.add_argument("--refs-per-speaker", type=int, default=1)
    p.add_argument("--tau", type=float, default=0.3)
    p.add_argument("--segment-seconds", type=float)
    p.add_argument("--normalize-refs-peak", type=float)
    p.add_argument("--demo", action="store_true")
    args = p.parse_args()
    if args.demo:
        demo()
    else:
        main(args.limit_sources, args.limit_speakers, args.out, args.refs_per_speaker, args.tau, args.segment_seconds, args.normalize_refs_peak)
