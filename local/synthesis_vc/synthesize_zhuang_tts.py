from pathlib import Path
import re

import torch
from openpyxl import Workbook, load_workbook
from scipy.io.wavfile import write
from transformers import VitsModel, AutoTokenizer


SRC = Path("壮语文本数据集样例.xlsx")
OUT = Path("outputs/mms-tts-zyb")
MODEL = "facebook/mms-tts-zyb"
SKIP_HEADERS = {
    "ID", "id", "Language", "language", "Domain", "domain", "Category", "category",
    "数据来源", "polarity", "secondary", "intent", "emotion", "culture_tag",
    "safety_tag", "role_id", "role_name", "至少3轮，4.5非必须",
}
LATIN = re.compile(r"[A-Za-z]")
ALLOWED = set("abcdefghijklmnopqrstuvwxyz '-—")


def clean(text: str) -> str:
    text = text.lower().replace("’", "'").replace("‘", "'")
    return re.sub(r"\s+", " ", "".join(ch if ch in ALLOWED else " " for ch in text)).strip()


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    device = "cuda:7" if torch.cuda.is_available() else "cpu"
    tokenizer = AutoTokenizer.from_pretrained(MODEL, local_files_only=True)
    model = VitsModel.from_pretrained(MODEL, local_files_only=True).to(device).eval()

    wb = load_workbook(SRC, read_only=True, data_only=True)
    rows = []
    for ws in wb.worksheets:
        data = list(ws.iter_rows(values_only=True))
        headers = [str(x).strip() if x is not None else "" for x in data[0]]
        for ridx, row in enumerate(data[1:], start=2):
            for cidx, value in enumerate(row, start=1):
                header = headers[cidx - 1] if cidx <= len(headers) else ""
                if header in SKIP_HEADERS or not isinstance(value, str) or not LATIN.search(value):
                    continue
                normalized = clean(value)
                if not normalized:
                    continue
                stem = f"{ws.title}_r{ridx}_c{cidx}_{header or 'text'}"
                wav = OUT / f"{stem}.wav"
                inputs = tokenizer(normalized, return_tensors="pt").to(device)
                with torch.no_grad():
                    audio = model(**inputs).waveform[0].cpu().float().numpy()
                write(wav, model.config.sampling_rate, audio)
                rows.append([ws.title, ridx, cidx, header, str(wav), value, normalized])
                print(wav)

    mwb = Workbook()
    mws = mwb.active
    mws.title = "manifest"
    mws.append(["sheet", "row", "col", "field", "wav_path", "original_text", "tts_text"])
    for row in rows:
        mws.append(row)
    mwb.save(OUT / "manifest.xlsx")
    print(f"wrote {len(rows)} wav files and {OUT / 'manifest.xlsx'}")


if __name__ == "__main__":
    main()
